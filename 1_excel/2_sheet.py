from openpyxl import Workbook
wb = Workbook()

# < Create a sheet >

ws = wb.create_sheet() # Create a new sheet next to sheet1. "ws" will be this sheet, not the init sheet.
ws.title = "MySheet"
ws.sheet_properties.tabColor = "ff66ff" # worksheet tab, not cell

ws1 = wb.create_sheet("YourSheet")

ws2 = wb.create_sheet("NewSheet", 2) # (,on which index)



new_ws = wb["NewSheet"] # Access to the sheet with the sheet's name  in the form of a dict.

print(wb.sheetnames) # Check all sheets's name in the form of a list

# < Duplicate a sheet >
new_ws["A1"] = "Test" # Value of A1 cell will be "Test"
target = wb.copy_worksheet(new_ws) # (Which sheet)
target.title = "Copied Sheet"

# If you save it under the same name, the previous file disappears.
wb.save("sample.xlsx")