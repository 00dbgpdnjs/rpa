from openpyxl import load_workbook
wb = load_workbook("sample.xlsx") # Import the workbook
ws = wb.active

# Import cell data as much as you want.
for x in range(1, 11):
    for y in range(1, 11):
        print(ws.cell(row=x, column=y).value, end=" ") # Without end=" ", line-break for every value. 
    print() # line-break whenever moved on to the next row

# Import data when you don't know the number of cells.
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()