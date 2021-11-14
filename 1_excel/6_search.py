from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# Students with English scores of 70 or higher.
for row in ws.iter_rows(min_row=2):
    if int(row[1].value) > 70:
        print(row[0].value, "번 학생 영어 성적 우수")


# Change 영어 to 컴퓨터
for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "영어":
            cell.value = "컴퓨터"

wb.save("sample_modified.xlsx")