from openpyxl import load_workbook


# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# # Get the value immediately from cell obj
# # Print formula, not a result value.
# for row in ws.values:
#     for cell in row:
#         print(cell)

# -----------------------------------------------------

# 계산[evaluate]되지 않은 상태의 데이터는 None 이라고 표시
# openpyxl을 통해서 작성한 수식에 대해서는 수식이 그대로 들어가는 것이어서 실제로 계산되지 않음. 그래서 엑셀파일을 열고 저장하고 닫아야 계산된 데이트를 가져올 수 있음
wb = load_workbook("sample_formula.xlsx", data_only=True) # ❗❗ KEY CODE ❗❗
ws = wb.active

for row in ws.values:
    for cell in row:
        print(cell) # >> None (cell Which formula put in  before opening and saving)