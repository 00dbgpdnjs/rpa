from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 1. Put in scores

ws.append(("학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"))

scores = [ # This var is a list which have 10 tuples
    (1,10,8,5,14,26,12),
    (2,7,3,7,15,24,18),
    (3,9,5,8,8,12,4),
    (4,7,8,7,17,21,18),
    (5,7,8,7,16,25,15),
    (6,3,5,8,8,17,0),
    (7,4,9,10,16,27,18),
    (8,6,6,6,15,19,17),
    (9,10,10,9,19,30,19),
    (10,9,8,8,20,25,20)
]

for s in scores:
    ws.append(s)

# 2. Modify all scores of quiz2 to 10
for idx, cell in enumerate(ws["D"]):
    # Skip title
    if idx == 0:
        continue

    cell.value = 10

# 3. H열에 총점(SUM 이용), I열에 성적 정보 추가
ws["H1"] = "총점"
ws["I1"] = "성적"

for idx, score in enumerate(scores, start=2): # 2nd arg: 엑셀에서 2열부터 점수 데이터가 시작하기 때문에 idx와 숫자를 맞추기 위해. start=0이 디폴트이고 내 맘대로 설정가능. idx가 0부터 시작하는데 0과 1이 스킵된다는 의미가 아님!
    # print(idx, "  ", score)

    # 3-1. 총점
    # 3-1-1. evalute 된 데이터 -> 여기서 성적 처리 하기 위해
    sum_val = sum(score[1:]) - score[3] + 10 # 번호[0] 제외한 점수 다 더함 - 퀴즈2 점수 + 10 (퀴즈2)

    # 3-1-2. evalute 안 된 데이터 -> 실제 엘셀 파일에
    ws.cell(row=idx, column=8).value = "=SUM(B{}:G{})".format(idx, idx) # start=2를 넣은 이유 / where : cell(x, y) / SUM(B2:G2), SUM(B3:G3)..

    # 3-2. 성적 처리
    grade = None
    if sum_val >= 90:
        grade = "A"
    elif sum_val >= 80:
        grade = "B"
    elif sum_val >= 70:
        grade = "C"
    else:
        grade = "D"

    if score[1] < 5:
        grade = "F"

    ws.cell(row=idx, column=9).value = grade


wb.save("scores.xlsx")