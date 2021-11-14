from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

ws.column_dimensions["A"].width = 5 # The width of A열 will be 5.
ws.row_dimensions[1].height = 50 # 1st line

#------------------------------------------------------------------------

a1 = ws["A1"] # 번호
b1 = ws["B1"] # 영어
c1 = ws["C1"] # 수학

a1.font = Font(color="FF0000", italic=True, bold=True) # (red, lean,)
b1.font = Font(color="CC33FF", name="Arial", strike=True) # (,font,취소선)
c1.font = Font(color="0000FF", size=20, underline="single")



thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

a1.border = thin_border # Apply the var to border of a1 cell
b1.border = thin_border
c1.border = thin_border

#--------------------------------------------------------------------------

# If it's over 90, apply green.
for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal="center",  vertical="center")
        # Besides[In addition to] center,  left, right, top and bottom

        # Except to A열[번호열]
        if cell.column == 1: 
            continue 

        if  isinstance(cell.value, int) and cell.value > 90: # func: In case that cell's value is int
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid")
            cell.font = Font(color="FF0000")

#-----------------------------------------------------------------------

# Fix the frame (based on B2). 
# So A열 and Title line are fixed.
ws.freeze_panes = "B2"

wb.save("sample_style.xlsx")