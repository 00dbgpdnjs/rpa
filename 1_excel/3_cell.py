from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

# Put 1 in the A1 cell
ws["A1"] = 1 
c = ws.cell(column=3, row=1, value=10)


print(ws["A1"].value)
print(ws.cell(row=1, column=1).value)
# print(ws.cell(column=1, row=1).value)
print(c.value)

from random import *

# A1 ... J1, A2 ... J2, A3 ... J10
index = 1 
for x in range(1, 11):
    for y in range(1, 11):
        # ws.cell(row=x, column=y, value=randint(0, 100)) # random integer of 0 ~ 100
        ws.cell(row=x, column=y, value=index)
        index += 1

wb.save("sample.xlsx")