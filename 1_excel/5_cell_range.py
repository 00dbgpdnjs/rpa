from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

# append() : 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])

# << This section : [:] Set range of row or col >>


# col_B = ws["B"] # Get only the B[영어] col
# for cell in col_B:
#     print(cell.value)

# col_range = ws["B:C"] # 영어, 수학 col
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

# row_title = ws[1] # 1st row
# for cell in row_title:
#     print(cell.value) # >> 번호 영어 수학

# row_range = ws[2:6] # 2nd ~ 6th lines
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

# row_range = ws[2:ws.max_row] # 2nd ~ last line[last value]
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

from openpyxl.utils.cell import coordinate_from_string

# # Print coordinates
# row_range = ws[2:ws.max_row]
# for rows in row_range:
#     for cell in rows:
#         # 1st way : >> like A2
#         # print(cell.coordinate, end=" ")

#         # 2nd way : >> like (x, y)
#         xy = coordinate_from_string(cell.coordinate)
#         # print(xy, end=" ")

#         # 3rd way : print the coordinates separately.
#         print(xy[0], end="") # x-coordinate
#         print(xy[1], end=" ") # y-coordinate
#     print()

#------------------------------------------------
# << This section : tuple() >>

print(tuple(ws.rows)) # One tuple means one line

# Print all tuple one by one
for row in tuple(ws.rows):
    print(row)

# Print value of 영어 column
for row in tuple(ws.rows):
    print(row[1].value)


print(tuple(ws.columns)) # One tuple means one column

for column in tuple(ws.columns):
    print(column[0].value) # >> 번호 영어 수학

#---------------------------------------------
# << This section : iter_rows() and iter_cols() >>

for row in ws.iter_rows(): # func: all rows
    print(row) # >> one line
    print(row[1].value) # 영어 점수 one by one so When this loop is over, all English scores are printed.

for column in ws.iter_cols():
    print(column[0].value) # >> 번호 영어 수학

# 4 students's math socre
for row in ws.iter_rows(min_row=1, max_row=5):
    print(row[2].value)

# all students's eng and math score
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3): # "row" is a tuple; (eng, math)
    print(row[0].value, row[1].value) # (수학, 영어)

wb.save("sample.xlsx")